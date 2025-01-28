# !/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook


def compare():
    wb = load_workbook(filename="compare.xlsx")
    comparison = dict()

    ws108 = wb['108']
    for row in ws108.iter_rows(values_only=True):
        v = row[0]

        if v not in comparison.keys():
            comparison[v] = {
                '108': row[1],
                '110': None
            }
        else:
            comparison[v]['108'] = row[1]

    ws110 = wb['110']
    for row in ws110.iter_rows(values_only=True):
        v = row[0]

        if v not in comparison.keys():
            comparison[v] = {
                '108': None,
                '110': row[1]
            }
        else:
            comparison[v]['110'] = row[1]

    # print(comparison['ADcer'])

    cnt = 0
    for v, comp in comparison.items():
        if comp['108'] is not None and comp['108'] != comp['110']:
            print(f"{v} => {comp['108']}")
            cnt += 1

    print(cnt)


if __name__ == "__main__":
    compare()
